"""Sinh workbook mapping MAU - dung lam fixture test va lam mau cho DE.

  python tools/gold_review/tests/make_fixture_mapping.py
-> tests/fixtures/mapping/GOOD_FCT_Silver_to_Gold.xlsx

Chua du 2 block ma tool doc: JOIN SCHEMA (tieu chi X.8) va FIELD MAPPING (tieu chi 1.1).
"""
from __future__ import annotations

import os

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

NAVY, BLUE = "FF000080", "FF0069BA"
ARIAL = Font(name="Arial", size=11)
BOLD = Font(name="Arial", size=11, bold=True)
HEAD = Font(name="Arial", size=11, bold=True, color="FFFFFFFF")

JOIN_SCHEMA = [
    ("sts_hub_customer", "d", "BASE", "GROUP BY hk HAVING max_by(cdc_status, source_event_date)='D'", "loc ban ghi da xoa"),
    ("hub_customer", "h", "LEFT JOIN", "d.customer_hashkey IS NULL (anti-join)", "III.4.2.1"),
    ("sat_customer_classification", "s", "LEFT JOIN", "s.customer_hashkey = a.customer_hashkey", "max_by + cutoff <="),
    ("link_account_customer", "lc", "LEFT JOIN", "rut current: GROUP BY account_hashkey", "III.4.2.3"),
    ("sat_account_balance", "b", "LEFT JOIN", "b.account_hashkey = lc.account_hashkey", "bang transaction -> loc ="),
]
FIELDS = [
    ("CST_ID", "STRING", "1:1", "a.customer_hashkey", "hub_customer"),
    ("BAL_AMT_LCY", "DECIMAL(20,4)", "Aggregate", "COALESCE(SUM(b.balance),0)", "sat_account_balance"),
    ("CDR_DT_ID", "INT", "Constant", "CAST(:DATADT AS INT)", "param ngay chay"),
]
SQL_ROWS = [
    ("Code cu", "OCBRT.GOOD_FCT", "-- DataStage job GOOD_FCT_01 (tham khao)"),
    ("Code moi", "GOOD_FCT (Gold)", "-- xem tests/fixtures/good_silver.sql"),
]


def main() -> list:
    out_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fixtures", "mapping")
    os.makedirs(out_dir, exist_ok=True)
    return [_good_fct(out_dir), _upl_fct(out_dir)]


def _upl_fct(out_dir: str) -> str:
    """Workbook cho object doc bang UPLOAD/THU CONG: nguon duoc khai o cot NOTE.
    Dung de kiem chung ngoai le cua tieu chi X.9 (bang upload la nguon hop le)."""
    path = os.path.join(out_dir, "UPL_FCT_Silver_to_Gold.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "UPL_FCT"

    _title(ws, 1, "[ VIEW ]  GOLD UPL_FCT  —  view tren bang upload thu cong")
    _title(ws, 3, "JOIN SCHEMA - UPL_FCT")
    _header(ws, 4, ["#", "Table / View (nguon)", "Alias", "JOIN Type", "ON Condition / Ghi chu", "Note"])
    _row(ws, 5, [1, "tb_manual_upl", "u", "BASE", "",
                 "TABLE upload thu cong (khong qua ETL/Silver)"])

    _title(ws, 7, "FIELD MAPPING - UPL_FCT")
    _header(ws, 8, ["STT", "Fields", "Data Type", "Type", "Transform", "Source"])
    _row(ws, 9, [1, "CDR_DT_ID", "INT", "1:1", "u.CDR_DT_ID", "tb_manual_upl"])
    _row(ws, 10, [2, "RATE_VAL", "DECIMAL(20,4)", "1:1", "u.RATE_VAL", "tb_manual_upl"])
    for col, w in zip("ABCDEF", (6, 34, 16, 12, 52, 44)):
        ws.column_dimensions[col].width = w

    wb.save(path)
    return path


def _good_fct(out_dir: str) -> str:
    path = os.path.join(out_dir, "GOOD_FCT_Silver_to_Gold.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GOOD_FCT"

    _title(ws, 1, "[ BulkLoad ]  GOLD GOOD_FCT  —  fixture mau, dung pattern Technical_Document III.4.2")
    _title(ws, 3, "JOIN SCHEMA")
    _header(ws, 4, ["#", "Table / View (nguon)", "Alias", "JOIN Type", "ON Condition / Ghi chu", "Note"])
    for i, row in enumerate(JOIN_SCHEMA, 1):
        _row(ws, 4 + i, [i, *row])

    r0 = 5 + len(JOIN_SCHEMA) + 1
    _title(ws, r0, "FIELD MAPPING - GOOD_FCT")
    _header(ws, r0 + 1, ["STT", "Fields", "Data Type", "Type", "Transform", "Source"])
    for i, row in enumerate(FIELDS, 1):
        _row(ws, r0 + 1 + i, [i, *row])
    for col, w in zip("ABCDEF", (6, 34, 16, 12, 52, 30)):
        ws.column_dimensions[col].width = w

    sq = wb.create_sheet("Script")   # sheet dump code, tool bo qua khi doc dac ta
    _header(sq, 1, ["Type", "Object", "Script SQL"])
    for i, row in enumerate(SQL_ROWS, 2):
        _row(sq, i, list(row))
    for col, w in zip("ABC", (12, 26, 56)):
        sq.column_dimensions[col].width = w

    wb.save(path)
    return path


def _title(ws, row: int, text: str) -> None:
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Arial", size=11, bold=True, color=NAVY)


def _header(ws, row: int, values: list) -> None:
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = HEAD
        c.fill = PatternFill("solid", fgColor=BLUE)
        c.alignment = Alignment(vertical="center", wrap_text=True)


def _row(ws, row: int, values: list) -> None:
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = ARIAL
        c.alignment = Alignment(vertical="top", wrap_text=(i >= 5))


if __name__ == "__main__":
    for p in main():
        print("da tao:", p)
