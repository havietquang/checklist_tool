#!/usr/bin/env python
"""Bo sung danh sach model tu TAI LIEU MAPPING SILVER vao engine/known_models.json.

Vi sao can: khoa "models" trong known_models.json trich tu docs/datavault-model-dev_zonec.zip
(dbt project nhanh dev) - anh chup tai thoi diem lay zip. Bang moi cua Zone C da mapping/dev
xong nhung chua co trong zip do se bi rule X.10 bao "KHONG TON TAI trong Data Vault model"
oan (vd hub_accountbc_save, link_accountbc_save_account, link_accountbc_save_branch,
sat_accountbc_save - nguon SBV).

Script doc sheet 'Tracking Mapping&Dev Zone C' cua tai lieu mapping, lay moi bang raw_vault/
business_vault TRU cac dong Dev Status = 'Cancel' (da huy thi khong ton tai that), ghi vao
khoa RIENG "models_tai_lieu_mapping" - KHONG tron vao "models" de con phan biet duoc nguon.

Chay lai moi khi nhan tai lieu mapping ban moi:
    python tools/gold_review/update_known_models.py "OCB ZoneC  - Tài liệu mapping silver (3).xlsx"
"""
from __future__ import annotations

import json
import os
import re
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("Thieu thu vien: pip install openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))
KM = os.path.join(HERE, "engine", "known_models.json")
SHEET = "Tracking Mapping&Dev Zone C"
# Chi nhan ten bang Data Vault that su - bo qua dong ghi chu / ten nguon.
# 'non_his_sat_' phai dung truoc 'sat_': mot so sat da chuyen thanh non-his sat va duoc them
# tien to non_his_ vao ten (CR 20260805/20260812), thieu tien to nay thi ca nhom bi bo qua.
PREFIX = re.compile(r"^(non_his_sat_|sts_hub_|hub_|link_|effsat_|csat_|sat_|ref_|bridge_|pit_)\w+$", re.I)
SKIP_STATUS = {"cancel"}


def read_tracking(path: str) -> tuple[set, set, dict]:
    """Tra ve (bang con hieu luc, bang da Cancel, thong ke theo Dev Status).

    Cancel doc theo CA HAI cot 'Mapping Status' va 'Dev Status': nhom app_product bi huy
    ngay 20260807 chi duoc danh Cancel o cot Mapping Status con Dev Status van la 'Checking',
    neu chi doc Dev Status thi bo lot ca nhom - dung loi da phat hien khi review Batch 2.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if SHEET not in wb.sheetnames:
        sys.exit(f"Khong thay sheet {SHEET!r} trong {os.path.basename(path)}")
    rows = list(wb[SHEET].iter_rows(values_only=True))
    hdr = [str(c or "").strip().lower() for c in rows[0]]
    try:
        i_tbl, i_dev = hdr.index("table name"), hdr.index("dev status")
        i_map = hdr.index("mapping status")
    except ValueError:
        sys.exit(f"Sheet {SHEET!r} thieu cot 'Table Name' / 'Mapping Status' / 'Dev Status'")

    tables, cancelled, by_status = set(), set(), {}
    for r in rows[1:]:
        name = str(r[i_tbl] or "").strip().lower()
        status = str(r[i_dev] or "").strip()
        if not PREFIX.match(name):
            continue
        if status.lower() in SKIP_STATUS or str(r[i_map] or "").strip().lower() in SKIP_STATUS:
            cancelled.add(name)
            continue
        tables.add(name)
        by_status[status or "(trong)"] = by_status.get(status or "(trong)", 0) + 1
    return tables, cancelled, by_status


def main() -> int:
    if len(sys.argv) < 2:
        return print(__doc__) or 2
    src = sys.argv[1]
    if not os.path.exists(src):
        sys.exit(f"Khong thay file: {src}")

    tables, cancelled, by_status = read_tracking(src)
    with open(KM, encoding="utf-8") as fh:
        data = json.load(fh)
    from_dbt = set(data["models"])
    extra = sorted(tables - from_dbt)
    # Giu nguyen mien tru da chot voi OCB qua cac lan review truoc.
    exempt = sorted({t.lower() for t in data.get("models_cancel_mien_tru", [])})

    # Bang Cancel phai BIEN MAT khoi danh sach doi chieu, khong chi nam o "models_cancel":
    # "models" trich tu zip dbt van con file model cua chung (Silver chua go), de nguyen thi
    # X.10 coi la hop le. Tru cac bang OCB xac nhan van dung.
    # "models_doi_ten" khai TAY (ten cu -> ten moi), khong suy tu dong duoc tu Change Request.
    # Giu nguyen qua moi lan sinh lai, va cung phai go ten cu khoi danh sach doi chieu.
    doi_ten = {k.lower() for k in data.get("models_doi_ten", {})}
    bo = sorted(((cancelled - set(exempt)) | doi_ten) & from_dbt)
    from_dbt -= set(bo)
    data["models"] = sorted(from_dbt)
    data["_so_luong"] = len(from_dbt)
    extra = sorted((set(extra) - cancelled - doi_ten) | (set(exempt) & cancelled))

    data["models_tai_lieu_mapping"] = extra
    data["_nguon_tai_lieu_mapping"] = (
        f"{os.path.basename(src)} sheet '{SHEET}' - bang da mapping/dev cho Zone C nhung"
        " CHUA co trong zip dbt (bo qua dong Cancel)."
        " Sinh lai bang: python tools/gold_review/update_known_models.py <file mapping>")
    data["models_cancel"] = sorted(cancelled)
    data["models_cancel_mien_tru"] = exempt
    data["_nguon_models_cancel"] = (
        f"{os.path.basename(src)} sheet '{SHEET}' - bang co Mapping Status HOAC Dev Status"
        " = 'Cancel'. Rule X.10 bao FAIL khi script Gold doc phai cac bang nay."
        " Bang nao OCB xac nhan van dung thi them tay vao 'models_cancel_mien_tru'.")
    with open(KM, "w", encoding="utf-8") as fh:
        json.dump(data, fh, ensure_ascii=False, indent=2)

    print(f"Doc {len(tables)} bang con hieu luc tu sheet '{SHEET}' (theo Dev Status: {by_status})")
    print(f"  - da co trong zip dbt        : {len(tables & from_dbt)}")
    print(f"  - BO SUNG vao known_models   : {len(extra)}")
    for t in extra[:10]:
        print(f"      {t}")
    if len(extra) > 10:
        print(f"      ... con {len(extra) - 10} bang")
    print(f"\nBang CANCEL (Mapping Status hoac Dev Status): {len(cancelled)}"
          f" - mien tru {len(exempt)}: {exempt}")
    print(f"Bang DOI TEN / TACH BANG (khai tay 'models_doi_ten'): {len(doi_ten)}")
    print(f"  - GO khoi 'models' (con file trong zip dbt nhung da huy / da doi ten): {len(bo)}")
    for t in bo[:10]:
        print(f"      {t}")
    if len(bo) > 10:
        print(f"      ... con {len(bo) - 10} bang")
    print(f"Tong danh sach doi chieu: {len(from_dbt | set(extra))} bang -> {KM}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
