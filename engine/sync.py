"""DOI CHIEU DONG BO - cung 1 object nhung code nam o nhieu noi thi noi nao dang giu ban cu.

Van de: bang DUNG CHUNG (T24_CRB, CB_OU_DIM, BRANCH_LIST, HOLIDAY...) duoc dan lai SQL vao
sheet Script cua NHIEU workbook tieu thu. Khi sua code, nguoi lam chi sua workbook CHINH CHU
cua bang do (vd '043. OCB_GOLD_TCKH_T24_CRB_PHAT.xlsx') va file src/tckh/t24_crb.sql, con cac
ban dan kem trong workbook tieu thu (vd '030. ..._V_T24_CRB_FULLLIST_PHAT.xlsx') van la ban cu
-> nguoi doc thiet ke thay 2 phien ban khac nhau cua cung 1 bang ma khong biet ban nao dung.

run_check.py da canh bao 'LECH NOI DUNG' nhung khong chi ra LECH O DAU, va khong so voi code
that trong src/tckh/*.sql. Module nay lam not:

  BAN CHUAN   = ban trong workbook CHINH CHU (ten file mang dung ten object) - dung mapping.owns.
                Khong co workbook chinh chu thi lay src/tckh/<object>.sql; khong co ca hai thi
                KHONG ket luan, chi liet ke cac ban de nguoi chon.
  So o 2 muc  : bo comment + gop whitespace (lech NOI DUNG - phai sua) va so nguyen van
                (chi khac FORMAT - de nguoi quyet).
  Canh bao rieng: o Excel toi da 32767 ky tu -> script dai hon bi CAT CUT im lang khi dan vao
                1 o; ban 'chuan' luc do cung khong con day du.

Chi BAO CAO, khong tu ghi de workbook - de nguoi doi chieu roi tu quyet dinh sua.
"""
from __future__ import annotations

import difflib
import glob
import hashlib
import os
import re
from dataclasses import dataclass, field

import mapping

# Excel gioi han 32767 ky tu / o. Script dai hon phai cat thanh NHIEU DONG cung khoi
# (_scan_script_sheet noi lai) hoac ghi DUONG DAN file .sql. Dan het vao 1 o thi Excel
# cat cut khong bao gi -> ban trong workbook thieu doan cuoi ma nhin mat thuong khong thay.
EXCEL_CELL_MAX = 32767
# Sat nguong cung dang nghi (co the da bi cat roi them vai ky tu, hoac sap bi cat).
NEAR_LIMIT = EXCEL_CELL_MAX - 200

CANON_OWNER = "workbook chinh chu"
CANON_SRC = "file src/tckh"
CANON_NONE = "KHONG XAC DINH"

_COMMENT_LINE_RE = re.compile(r"--[^\n]*")
_COMMENT_BLOCK_RE = re.compile(r"/\*.*?\*/", re.S)


def norm_content(sql: str) -> str:
    """Chuan hoa de so NOI DUNG LOGIC: bo comment, gop whitespace, ha chu thuong.

    Hai ban chi khac thut le / xuong dong / cau comment tieng Viet la CUNG MOT code -
    khong nen bao dong nhu ban lech logic that. Dau ';' CUOI cung cung bi bo: co/khong co
    ';' o cuoi file khong doi gi ve logic, de nguyen thi AR_DIM bi bao 'LECH' chi vi
    file .sql ket thuc bang ';' ma o Excel thi khong."""
    s = (sql or "").replace("\r\n", "\n").replace("\r", "\n")
    s = _COMMENT_BLOCK_RE.sub(" ", s)
    s = _COMMENT_LINE_RE.sub(" ", s)
    s = re.sub(r"\s+", " ", s).strip().lower()
    return re.sub(r"[\s;]+$", "", s)


def fingerprint(sql: str) -> str:
    return hashlib.sha1(norm_content(sql).encode()).hexdigest()[:8]


# Loai lech
D_SAME = "khop"
D_FORMAT = "chi khac format/comment"
D_CONTENT = "LECH NOI DUNG"


@dataclass
class Place:
    """Mot noi dang giu 1 ban code cua object."""
    path: str
    sheet: str          # ten sheet, rong neu la file .sql
    sql: str

    @property
    def is_file(self) -> bool:
        return not self.sheet

    @property
    def label(self) -> str:
        base = os.path.basename(self.path)
        return base if self.is_file else f"{base} [{self.sheet}]"

    @property
    def nchars(self) -> int:
        return len(self.sql or "")

    @property
    def truncated(self) -> bool:
        """O Excel co dau hieu bi CAT CUT (chi ap dung cho ban dan trong workbook).

        Do dai NAM TRONG [NEAR_LIMIT, EXCEL_CELL_MAX] moi dang nghi. VUOT QUA
        EXCEL_CELL_MAX thi chac chan code da duoc cat thanh NHIEU O (mapping noi lai) nen
        khong he bi cat - bao 'cat cut' luc do la bao oan (vd 077. LOAN_SUMMARY_LIST
        38412 ky tu). Dung 32767 = gan nhu chac chan bi cat."""
        return (not self.is_file) and NEAR_LIMIT <= self.nchars <= EXCEL_CELL_MAX

    def diff_kind(self, other: "Place") -> str:
        if norm_content(self.sql) != norm_content(other.sql):
            return D_CONTENT
        return D_SAME if self.sql == other.sql else D_FORMAT


@dataclass
class ObjSync:
    """Ket qua doi chieu cua 1 object."""
    object: str
    canon: Place | None = None
    canon_why: str = CANON_NONE
    same: list = field(default_factory=list)       # list[Place] khop ban chuan
    fmt_only: list = field(default_factory=list)   # list[Place] chi khac format
    stale: list = field(default_factory=list)      # list[Place] LECH NOI DUNG
    truncated: list = field(default_factory=list)  # list[Place] o Excel bi cat cut
    others: list = field(default_factory=list)     # list[Place] khi khong xac dinh ban chuan
    # Object khong co workbook CHINH CHU nao (bang dung chung chi duoc dan kem o workbook
    # tieu thu). Ghi nhan rieng vi day la thieu SOT TAI LIEU, khong phai lech code - khi
    # cac ban dan kem van khop nhau thi khong co gi phai sua ngay.
    no_owner: bool = False

    @property
    def n_places(self) -> int:
        return len(self.same) + len(self.fmt_only) + len(self.stale) + len(self.others) + (
            1 if self.canon else 0)

    @property
    def status(self) -> str:
        if self.canon is None:
            # Khong co ban chuan nhung moi ban deu khop -> chua can ai quyet gi.
            return ("dong bo" if len({norm_content(p.sql) for p in self.others}) <= 1
                    else "KHONG RO BAN CHUAN")
        if self.stale:
            return "LECH"
        if self.truncated:
            return "BI CAT CUT"
        if self.fmt_only:
            return "khac format"
        return "dong bo"

    @property
    def severity(self) -> int:
        """De sap xep: 0 = nang nhat."""
        return {"LECH": 0, "KHONG RO BAN CHUAN": 1, "BI CAT CUT": 2,
                "khac format": 3, "dong bo": 4}[self.status]

    def diff(self, place: Place, ctx_lines: int = 1, max_lines: int = 14) -> list:
        """Diff rut gon giua ban chuan va `place`, da bo comment/whitespace nhieu."""
        if not self.canon:
            return []
        left = _diff_lines(self.canon.sql)
        right = _diff_lines(place.sql)
        out = []
        for line in difflib.unified_diff(left, right, "ban chuan", "ban dang xet",
                                         n=ctx_lines, lineterm=""):
            if line.startswith(("---", "+++")):
                continue
            out.append(line)
            if len(out) >= max_lines:
                out.append(f"... (con nua, chay --sync-object {self.object} de xem day du)")
                break
        return out


def _diff_lines(sql: str) -> list:
    """Tach dong de diff: bo comment va thut le nen diff chi con khac biet THAT."""
    lines = []
    for raw in _COMMENT_BLOCK_RE.sub(" ", sql or "").split("\n"):
        s = _COMMENT_LINE_RE.sub("", raw).strip()
        if s:
            lines.append(re.sub(r"\s+", " ", s))
    return lines


def collect(mapping_dirs: list, src_dirs: list) -> dict:
    """{khoa_object: [Place, ...]} gop tu sheet Script cua workbook VA file .sql trong src.

    Dung lai mapping.discover_scripts() (da xu ly: cell cat nhieu dong, o ghi duong dan
    file .sql, loc dong Type='Code moi') thay vi doc lai workbook o day."""
    variants: dict = {}
    mapping.discover_scripts([d for d in mapping_dirs if os.path.isdir(d)], None, variants)
    places: dict[str, list] = {
        key: [Place(p, sheet, sql) for p, sheet, sql in cands]
        for key, cands in variants.items()
    }
    for d in src_dirs:
        for p in sorted(glob.glob(os.path.join(d, "*.sql"))):
            key = mapping.clean_key(os.path.splitext(os.path.basename(p))[0])
            if not key:
                continue
            with open(p, encoding="utf-8", errors="replace") as fh:
                places.setdefault(key, []).append(Place(p, "", fh.read()))
    return places


def analyse(places: dict) -> list:
    """[ObjSync] cho moi object co TU 2 NOI tro len giu code (1 noi thi khong the lech)."""
    out = []
    for key, group in places.items():
        if len(group) < 2:
            continue
        res = ObjSync(key)
        canon = _pick_canonical(key, group)
        if canon is None:
            res.no_owner = not any(not p.is_file and mapping.owns(p.path, key) for p in group)
            res.others = sorted(group, key=lambda p: p.label)
            res.truncated = [p for p in group if p.truncated]
            out.append(res)
            continue
        res.canon, res.canon_why = canon
        for p in group:
            if p is res.canon:
                continue
            bucket = {D_SAME: res.same, D_FORMAT: res.fmt_only, D_CONTENT: res.stale}
            bucket[res.canon.diff_kind(p)].append(p)
        res.truncated = [p for p in group if p.truncated]
        for lst in (res.same, res.fmt_only, res.stale, res.others):
            lst.sort(key=lambda p: p.label)
        out.append(res)
    return sorted(out, key=lambda r: (r.severity, r.object))


def _pick_canonical(key: str, group: list) -> tuple | None:
    """(Place, ly_do) cua ban CHUAN, hoac None khi khong xac dinh duoc.

    Uu tien workbook CHINH CHU (dung mapping.owns - cung luat ma run_check dang cham diem,
    nen ban duoc coi la chuan o day DUNG la ban da duoc cham 'pass'). Khong co thi lay file
    src/tckh/<object>.sql. Nhieu workbook cung so huu -> khong doan, de nguoi chon."""
    owned = [p for p in group if not p.is_file and mapping.owns(p.path, key)]
    if len({os.path.abspath(p.path) for p in owned}) == 1:
        return owned[0], CANON_OWNER
    if owned:
        return None                       # >1 workbook chinh chu: khong the tu chon
    files = [p for p in group if p.is_file]
    if len(files) == 1:
        return files[0], CANON_SRC
    return None


# ------------------------------------------------------------------ in ket qua
def print_report(results: list, show_diff: bool = True) -> None:
    lech = [r for r in results if r.status == "LECH"]
    unknown = [r for r in results if r.status == "KHONG RO BAN CHUAN"]
    cut = [r for r in results if r.status == "BI CAT CUT"]
    fmt = [r for r in results if r.status == "khac format"]
    ok = [r for r in results if r.status == "dong bo"]

    print(f"\n{'=' * 100}")
    print("DOI CHIEU DONG BO - cung 1 object, code nam o nhieu noi")
    print(f"{'=' * 100}")
    # Dem NOI bi cat cut tren TOAN BO ket qua: o cat cut hay di kem object da 'LECH' (status
    # LECH uu tien cao hon) nen chi dem nhom `cut` se ra 0 va nguoi doc tuong khong co.
    n_cut = sum(len(r.truncated) for r in results)
    print(f"  {len(results)} object co code o >= 2 noi:  {len(lech)} LECH NOI DUNG"
          f" | {len(unknown)} khong ro ban chuan | {len(fmt)} chi khac format"
          f" | {len(ok)} dong bo")
    if n_cut:
        where = (f" - {len(cut)} object chi bi loi nay, so con lai nam trong object da LECH"
                 if cut else " - deu nam trong object da LECH o tren")
        print(f"  {n_cut} o Excel co dau hieu BI CAT CUT o gioi han {EXCEL_CELL_MAX} ky tu"
              f"{where}")
    print("  Ban CHUAN = ban trong workbook CHINH CHU - ten file mang dung ten object"
          " (vd '043. OCB_GOLD_TCKH_T24_CRB_PHAT.xlsx' la chinh chu cua T24_CRB).")

    for r in lech:
        print(f"\n[!] LECH  {r.object}   ({r.n_places} noi giu code)")
        print(f"      ban chuan : {r.canon.label}   ({r.canon.nchars} ky tu, {r.canon_why})")
        for p in r.stale:
            tag = "  <-- O EXCEL BI CAT CUT" if p.truncated else ""
            print(f"      LECH      : {p.label}   ({p.nchars} ky tu){tag}")
            if show_diff:
                for line in r.diff(p):
                    print(f"                  {line}")
        for p in r.fmt_only:
            print(f"      khac format: {p.label}")
        for p in r.same:
            print(f"      khop       : {p.label}")

    for r in unknown:
        print(f"\n[?] KHONG RO BAN CHUAN  {r.object}   ({r.n_places} noi giu code,"
              f" {len({fingerprint(p.sql) for p in r.others})} phien ban khac nhau)")
        why = ("khong noi nao mang dung ten object -> object nay CHUA CO workbook chinh chu"
               if r.no_owner else "nhieu workbook cung mang ten object nay")
        print(f"      ly do: {why}; tool khong tu chon, nguoi phai chi ro ban dung")
        for p in r.others:
            print(f"      - {fingerprint(p.sql)}  {p.nchars:>7} ky tu  {p.label}")

    for r in cut:
        print(f"\n[!] BI CAT CUT  {r.object}   ban chuan {r.canon.label}")
        for p in r.truncated:
            print(f"      - {p.label}   {p.nchars} ky tu >= gioi han o Excel {EXCEL_CELL_MAX}")
        print("      -> Excel cat cut IM LANG. Sua: cat code thanh nhieu dong cung khoi"
              " (de trong cot View/Table o cac dong sau), hoac ghi DUONG DAN file .sql"
              " vao o thay vi dan code.")

    if fmt:
        print(f"\n[i] {len(fmt)} object chi khac FORMAT/COMMENT (noi dung logic giong nhau)"
              " - khong bat buoc sua:")
        for r in fmt:
            print(f"      {r.object:<30} chuan {os.path.basename(r.canon.path)}"
                  f"  vs  {', '.join(p.label for p in r.fmt_only)}")

    orphan = [r for r in ok if r.no_owner]
    if orphan:
        print(f"\n[i] {len(orphan)} object CHUA CO workbook chinh chu (chi duoc dan kem trong"
              " workbook cua object khac) - cac ban dan kem dang khop nhau nen chua phai sua,"
              "\n    nhung khi sua code se khong biet phai sua o dau:")
        for r in orphan:
            print(f"      {r.object:<30} {r.n_places} noi: "
                  f"{', '.join(os.path.basename(p.path) for p in r.others)}")

    print(f"\n{'-' * 100}")
    if lech or unknown or cut:
        print("Huong sua (uu tien): moi object chi nen co code o DUNG 1 noi - workbook chinh chu"
              " cua no.\n  O workbook tieu thu: xoa dong 'Code moi' dan kem, chi khai ten bang"
              " o block JOIN SCHEMA;\n  hoac thay code bang DUONG DAN file .sql"
              " (vd 'src/tckh/t24_crb.sql') - tool tu doc file nen khong bao gio lech nua.")
    else:
        print("Tat ca object da dong bo.")


def print_object(results: list, obj: str) -> int:
    """In diff DAY DU cua 1 object.

    Exit code: 0 = da dong bo | 1 = con lech / khong ro ban chuan | 2 = khong thay object."""
    key = mapping.clean_key(obj)
    match = next((r for r in results if r.object == key), None)
    if match is None:
        print(f"Khong thay object '{obj}' co code o >= 2 noi.")
        return 2
    print(f"\n{'=' * 100}\n{match.object}   trang thai: {match.status}\n{'=' * 100}")
    if match.canon:
        print(f"ban chuan: {match.canon.label}   ({match.canon.nchars} ky tu, {match.canon_why})")
    for p in match.stale + match.fmt_only:
        print(f"\n--- vs {p.label}   ({p.nchars} ky tu, {match.canon.diff_kind(p)})")
        for line in match.diff(p, ctx_lines=2, max_lines=10 ** 6):
            print(f"  {line}")
    for p in match.others:
        print(f"\n--- ban tai {p.label}   ({p.nchars} ky tu, dau van tay {fingerprint(p.sql)})")
    if not (match.stale or match.fmt_only or match.others):
        print("Cac noi con lai khop y nguyen ban chuan.")
    return 1 if match.status in ("LECH", "KHONG RO BAN CHUAN") else 0


# ------------------------------------------------------------------ xuat Excel
HDRS = ["STT", "OBJECT", "TRANG THAI", "BAN CHUAN", "KY TU CHUAN", "CAN CU CHUAN",
        "NOI DANG XET", "KY TU", "LOAI LECH", "CAT CUT?", "DIFF TOM TAT"]


def write_excel(results: list, out_path: str) -> str:
    """Xuat bang doi chieu ra Excel de gui doi soat. Tra ve duong dan file."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Doi chieu dong bo"
    head_fill = PatternFill("solid", fgColor="000080")
    for i, h in enumerate(HDRS, 1):
        c = ws.cell(1, i, h)
        c.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        c.fill = head_fill
        c.alignment = Alignment(vertical="center", wrap_text=True)

    bad_fill = PatternFill("solid", fgColor="FFC7CE")
    warn_fill = PatternFill("solid", fgColor="FFEB9C")
    row, stt = 2, 0
    for r in results:
        stt += 1
        targets = ([(p, r.canon.diff_kind(p) if r.canon else "") for p in r.stale + r.fmt_only]
                   + [(p, "khong ro ban chuan") for p in r.others])
        if not targets:
            targets = [(None, "")]
        for p, kind in targets:
            vals = [
                stt, r.object, r.status,
                r.canon.label if r.canon else "", r.canon.nchars if r.canon else "",
                r.canon_why,
                p.label if p else "", p.nchars if p else "", kind,
                "CO" if (p and p.truncated) else "",
                "\n".join(r.diff(p)) if (p and r.canon and kind == D_CONTENT) else "",
            ]
            for i, v in enumerate(vals, 1):
                c = ws.cell(row, i, v)
                c.font = Font(name="Arial", size=11)
                c.alignment = Alignment(vertical="top", wrap_text=(i == len(HDRS)))
                if r.status in ("LECH", "KHONG RO BAN CHUAN"):
                    c.fill = bad_fill
                elif r.status == "BI CAT CUT":
                    c.fill = warn_fill
            row += 1

    for col, width in zip("ABCDEFGHIJK",
                          (5, 30, 20, 46, 12, 18, 46, 10, 22, 10, 90)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "B2"
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)
    return out_path
