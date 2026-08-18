"""Xuat ket qua ra ban Excel MOI dua tren workbook checklist goc (file goc khong doi)."""
from __future__ import annotations

import os
import shutil
from copy import copy
from datetime import date

import openpyxl
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill

import rule_text
from rules import BLOCKING, FAIL, NA, PASS, RULES, WARN

SHEET_REVIEW = "Review theo bảng"
SHEET_DETAIL = "Auto-check chi tiết"

# rule id -> cot tren sheet 'Review theo bảng'
COL = {
    # Tieu chi 1.3 (doi soat so dong & SUM) da bo han khoi checklist, nhom 1 danh so
    # lai con 1.1 / 1.2 (1.2 = danh sach bang khop thiet ke, truoc day mang ma 1.4).
    "1.1": "F", "1.2": "G",
    "2.1": "H", "2.2": "I", "2.3": "J", "2.4": "K", "2.5": "L", "2.6": "M",
    "2.7": "N", "2.8": "O", "2.9": "P", "2.10": "Q", "2.11": "R", "2.12": "S",
    "3.1": "T", "3.2": "U", "3.3": "V", "3.4": "W", "3.5": "X", "3.6": "Y",
}
VERDICT_COL = "AE"                    # cot 'Kết luận' - vốn là công thức =IF(ty_le<=nguong,...)
NOTE_COL = "AF"                       # cot 'Ghi rõ: mã tiêu chí – file script – dòng/CTE'
# Cot THEM MOI (checklist goc het o AF): canh bao dong bo tai lieu. Tach rieng chu khong
# gop vao AF de de nhin - AF da rat dai vi chua giai trinh tung tieu chi FAIL/WARN, nhet
# them canh bao dong bo vao thi phai keo het o moi thay.
SYNC_COL = "AG"
SYNC_HDR = "Đồng bộ tài liệu\n(script khác bản chính)"
FIRST_ROW, LAST_ROW = 3, 200      # row 3 = dong du lieu dau tien (khop CF Pass/Fail co san tu F3)
HDR_ROW = 2                       # dong tieu de cua sheet 'Review theo bảng'

NAVY = "FF000080"
BLUE = "FF0069BA"
GREEN = "FF1FBA00"
WARN_BG = "FFFFF2CC"          # nen vang nhat cho o WARN (canh bao, khong tinh loi)
WARN_FG = "FF9C6500"          # chu nau vang cho o WARN
BLOCK_BG = "FFFFC7CE"         # nen do nhat cho o 'Không đạt' bi rule chan ep xuong
BLOCK_FG = "FF9C0006"         # chu do dam
ARIAL = Font(name="Arial", size=11)
HEAD = Font(name="Arial", size=11, bold=True, color="FFFFFFFF")


# Sheet ket qua chi nhan 2 gia tri Pass / Fail:
#   PASS   -> Pass
#   N-A    -> Pass  (khong ap dung thi khong co gi sai; cong thuc Excel chi dem "Fail")
#   WARN   -> WARN  (canh bao rieng, to vang nhat - KHONG tinh loi, xem giai thich duoi)
#   FAIL   -> Fail  (chac chan sai, moi tinh loi)
#
# WARN khong tinh loi vi day la cac dau hieu "kem toi uu / nen xem lai" chu khong phai
# sai ket qua (vd doc lai bang calendar o 2 CTE, CTE dung lai chua CACHE). De WARN =
# Fail thi chi mot goi y toi uu cung day ca bang xuong "Khong dat" (nhom 2 = 80%), lam
# nhieu ket luan gui OCB.
# Ghi han chu 'WARN' (khong phai 'Pass') de nhin la thay ngay cho can toi uu. Cong thuc
# ty le loi cua checklist dem COUNTIF(...,"Fail") nen 'WARN' khong bi tinh la loi - khop
# dung voi score() ben run_check.py.
CELL = {PASS: "Pass", NA: "Pass", WARN: "WARN", FAIL: "Fail"}


def _cell_value(status: str) -> str:
    return CELL.get(status, "Fail")


def _available_path(path: str) -> str:
    """Moi lan chay ghi ra mot BAN MOI, khong de len ban cu.

    Lan dau ra '<ten>.xlsx', cac lan sau tu tang '<ten>_v2.xlsx', '_v3'... de con giu lai
    ket qua cu ma doi chieu, va cung khong bao gio vo PermissionError khi ban cu dang mo
    trong Excel (bo qua luon ca file dang bi khoa)."""
    if not os.path.exists(path):
        return path
    base, ext = os.path.splitext(path)
    i = 2
    while os.path.exists(f"{base}_v{i}{ext}"):     # chua ton tai => chac chan ghi duoc
        i += 1
    return f"{base}_v{i}{ext}"


def write(results: list, template: str, out_path: str, batch: str, pic: str,
          sync_notes: dict | None = None) -> str:
    """results: list[(ctx, {rule_id: Finding})]. Tra ve duong dan THAT SU da ghi (co
    the khac out_path neu file dich dang bi khoa - xem _available_path).

    `sync_notes`: {ten_workbook: [dong ghi chu]} - canh bao DONG BO TAI LIEU (workbook nay
    dang giu ban CU cua bang dung chung nao). Ghi vao cot Ghi chu cua dung dong object do,
    vi canh bao chi in ra console thi mo file Excel len khong thay gi."""
    out_path = _available_path(out_path)
    shutil.copyfile(template, out_path)
    wb = openpyxl.load_workbook(out_path)
    ws = wb[SHEET_REVIEW]

    dup = {c.label for i, (c, _) in enumerate(results)
           if any(o.label == c.label for j, (o, _) in enumerate(results) if j != i)}

    row = FIRST_ROW
    for ctx, findings in results:
        while row <= LAST_ROW and ws[f"B{row}"].value:
            row += 1
        if row > LAST_ROW:
            print(f"  [!] het cho tren sheet Review (toi da dong {LAST_ROW}), bo qua {ctx.label}")
            break

        if ctx.stt is not None:
            ws[f"A{row}"] = ctx.stt          # khop dung STT cua workbook thiet ke, khong danh lai
        ws[f"B{row}"] = (f"{ctx.label} ({os.path.basename(ctx.path)})"
                         if ctx.label in dup else ctx.label)
        ws[f"C{row}"] = batch
        # --pic tren dong lenh (neu co) ghi de cho ca batch; khong truyen thi lay PIC
        # rieng cua tung object (boc tu doan cuoi ten workbook thiet ke).
        ws[f"D{row}"] = pic or ctx.pic
        ws[f"E{row}"] = date.today().isoformat()
        for c in "BCDE":
            ws[f"{c}{row}"].font = ARIAL

        for rid, col in COL.items():
            f = findings.get(rid)
            if f is None:
                continue
            cell = ws[f"{col}{row}"]
            cell.value = _cell_value(f.status)
            cell.font = ARIAL

        # Rule CHAN: nhom X khong co cot tren sheet nen khong vao cong thuc ty le loi. Ghi
        # DE cong thuc o cot Ket luan de FAIL X.10 khong bi ket luan la "Đạt".
        blocked = sorted(rid for rid in BLOCKING
                         if (findings.get(rid) is not None
                             and findings[rid].status == FAIL))
        if blocked:
            cell = ws[f"{VERDICT_COL}{row}"]
            cell.value = "Không đạt"
            cell.font = Font(name="Arial", size=11, bold=True, color=BLOCK_FG)
            cell.fill = PatternFill("solid", start_color=BLOCK_BG, end_color=BLOCK_BG)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws[f"{NOTE_COL}{row}"] = _note(findings)
        ws[f"{NOTE_COL}{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws[f"{NOTE_COL}{row}"].font = ARIAL

        lines = _sync_lines(ctx, sync_notes)
        cell = ws[f"{SYNC_COL}{row}"]
        cell.value = "\n".join(lines) if lines else ""
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.font = Font(name="Arial", size=11, color=NAVY) if lines else ARIAL
        row += 1

    _setup_sync_col(ws, row)
    _setup_warn_format(ws)
    _detail_sheet(wb, results)
    wb.save(out_path)
    return out_path


def _setup_sync_col(ws, after_row: int) -> None:
    """Ghi tieu de + do rong cho cot SYNC_COL (cot them moi, checklist goc khong co).

    Style tieu de sao y cot AF ke ben de nhin lien mach voi bang goc; ke luon vien cho cac
    o trong tu dong du lieu cuoi tro len de cot khong bi 'ho' giua bang."""
    src = ws[f"{NOTE_COL}{HDR_ROW}"]
    dst = ws[f"{SYNC_COL}{HDR_ROW}"]
    dst.value = SYNC_HDR
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.column_dimensions[SYNC_COL].width = 52
    for r in range(FIRST_ROW, after_row):
        ws[f"{SYNC_COL}{r}"].border = copy(ws[f"{NOTE_COL}{r}"].border)


def _setup_warn_format(ws) -> None:
    """Cho o co gia tri 'WARN' hien mau VANG NHAT, va them 'WARN' vao dropdown.

    Dung DINH DANG CO DIEU KIEN chu khong phai cell.fill: trong Excel CF luon de len
    dinh dang truc tiep, ma file checklist da co san CF to Pass/Fail/N-A tren ca vung
    F3:Y60 -> to fill tay se khong hien gi. Rule nay dat priority = 1 cho chac.
    Them 'WARN' vao danh sach dropdown de o khong bi coi la du lieu khong hop le khi
    nguoi review mo file ra sua tay."""
    rng = f"F{FIRST_ROW - 1}:Y{LAST_ROW}"
    warn_fill = PatternFill("solid", start_color=WARN_BG, end_color=WARN_BG)
    rule = CellIsRule(operator="equal", formula=['"WARN"'], fill=warn_fill,
                      font=Font(color=WARN_FG, bold=True))
    ws.conditional_formatting.add(rng, rule)
    others = [r for cf in ws.conditional_formatting for r in cf.rules if r is not rule]
    rule.priority = 1
    for i, r in enumerate(others, start=2):
        r.priority = i

    for dv in ws.data_validations.dataValidation:
        if dv.type == "list" and dv.formula1 and "Pass" in dv.formula1 and "WARN" not in dv.formula1:
            dv.formula1 = dv.formula1.rstrip('"') + ',WARN"'


def _sort_ids(ids) -> list:
    return sorted(ids, key=_sort_key)


# Nhom X (X.1..X.9) la tieu chi BO SUNG cua Raffles, KHONG co trong checklist OCB va
# khong co cot tren sheet ket qua. Mac dinh KHONG dua vao file gui OCB de nguoi doc khong
# tuong day la tieu chi hai ben da thong nhat.
# RIENG 3 tieu chi duoi co TRICH DAN Technical Document ZoneC v1.1 (kiem chung duoc) nen
# van dua vao file, kem cau trich dan de minh bach la tieu chi bo sung, khong phai
# checklist. Cac tieu chi X con lai (X.2/X.3/X.4/X.5/X.8/X.9) khong co nguon tai lieu ->
# chi in ra man hinh cho doi DE tu xem.
X_DOCUMENTED = {
    "X.1": "Technical Document ZoneC v1.1, mục III.4.2.3",
    "X.6": "Technical Document ZoneC v1.1, mục III.4.2.2 nguyên tắc 1",
    "X.7": "Technical Document ZoneC v1.1, mục Cấu hình Pipeline / Cú pháp chung",
    "X.10": "đối chiếu trực tiếp với docs/datavault-model-dev_zonec.zip (models/raw_vault, models/business_vault) - danh sách bảng model dbt thật; và trạng thái Cancel ở tài liệu mapping Silver (sheet 'Tracking Mapping&Dev Zone C', cột Mapping Status + Dev Status)",
    "X.11": "đối chiếu trực tiếp với docs/datavault-model-dev_zonec.zip (models/raw_vault/satellite) - cột list_cols/raw_sql cua model dbt that",
    "X.12": "satellite multiactive (unique_key có ma_key trong docs/datavault-model-dev_zonec.zip): 1 hashkey nhiều dòng trong cùng ngày nên GROUP BY phải kèm ma_key, cột tiền phải SUM thay vì max_by",
    "X.13": "bảng transaction đã lọc source_event_date = 1 ngày thì max_by không chọn gì, bỏ được",
    "X.14": "PIT có khai sts_hub_table trong docs/datavault-model-dev_zonec.zip đã tự loại bản ghi cdc_status='D', join thêm hub/*_active là thừa",
    "X.15": "PIT đã có sẵn cột <sat>_src_ev_dt nên join thẳng satellite theo cột đó, không cần bọc thêm CTE/subquery",
}


def in_checklist(rid: str) -> bool:
    """Tieu chi co duoc ghi vao file Excel gui OCB khong (xem X_DOCUMENTED)."""
    return not rid.startswith("X.") or rid in X_DOCUMENTED


def _origin_tag(rid: str) -> str:
    """Nhan nguon cho tieu chi bo sung, de khong bi nham la tieu chi checklist OCB."""
    return f"   [bổ sung, không tính điểm — {X_DOCUMENTED[rid]}]" if rid in X_DOCUMENTED else ""


def _sync_lines(ctx, sync_notes: dict | None) -> list:
    """Cac dong canh bao dong bo cua workbook sinh ra `ctx`.

    ctx.path o che do doc workbook co dang '<ten workbook>.xlsx [<sheet>: <OBJ>]' -> cat
    lay ten workbook truoc dau ' ['. Che do --from-sql thi ctx.path la file .sql, khong
    khop khoa nao nen khong ghi gi (dung: luc do khong doc workbook de ma so)."""
    if not sync_notes:
        return []
    return sync_notes.get(os.path.basename(ctx.path.split(" [")[0]), [])


def _note(findings: dict) -> str:
    """Note chi tiết cho cột 'Ghi chú / Bằng chứng lỗi': mỗi tiêu chí 1 khối gồm
    tên tiêu chí -> lý do (kèm số dòng code) -> cách sửa.
    Chi ghi tieu chi CO trong checklist OCB (xem in_checklist)."""
    failed = _sort_ids([rid for rid, f in findings.items()
                        if f.status == FAIL and in_checklist(rid)])
    warned = _sort_ids([rid for rid, f in findings.items()
                        if f.status == WARN and in_checklist(rid)])
    blocks = []

    if failed:
        blocks.append(f"╔══ FAIL: {len(failed)} tiêu chí — {', '.join(failed)}")
        # Object khong doc duoc code (loi cu phap / thieu CREATE) thi MOI tieu chi deu
        # Fail voi CUNG mot ly do -> ghi ly do do 1 lan, khong lap 20 khoi giong het nhau.
        same = {tuple(findings[rid].evidence) for rid in failed}
        if len(same) == 1 and len(failed) > 3:
            for e in same.pop():
                blocks.append(f"    Lý do chung: {e}")
        else:
            for rid in failed:
                f = findings[rid]
                blocks.append(f"\n▼ {rid} — {rule_text.title(rid, RULES[rid].title)}{_origin_tag(rid)}")
                for e in f.evidence:
                    blocks.append(f"   • Lý do: {e}")
                if rule_text.fix(rid):
                    blocks.append(f"   → Cách sửa: {rule_text.fix(rid)}")

    # WARN va CHUA CO SO LIEU khong phai loi, nhung VAN phai ghi ro LY DO tung tieu chi -
    # chi liet ke ma khong (2.7, 2.8...) thi nguoi doc khong biet vuong o dau. Chi bo phan
    # "cach sua" dai dong va cau giai thich chung chung lap lai o moi dong.
    if warned:
        blocks.append(f"\n╔══ WARN: {len(warned)} tiêu chí — {', '.join(warned)}")
        for rid in warned:
            blocks.append(f"\n▼ {rid} — {rule_text.title(rid, RULES[rid].title)}{_origin_tag(rid)}")
            for e in findings[rid].evidence:
                blocks.append(f"   • {e}")

    # Tieu chi N-A (script khong dung sts/satellite/CASE/window... nen khong co gi de sai)
    # KHONG ghi vao note: note chi de giai trinh cho tieu chi FAIL hoac WARN.
    # Chi tiet day du van con o sheet 'Auto-check chi tiet'.
    # Canh bao lech dong bo nam o cot SYNC_COL rieng, KHONG gop vao day: do khong phai loi
    # cua chinh object nay (code cua no van dung) nen khong the de lan giua cac khoi FAIL.
    if not blocks:
        return "Đạt toàn bộ tiêu chí."
    return "\n".join(blocks).lstrip("\n")      # bo dong trong dau o khi khong co muc FAIL


def _detail_sheet(wb, results: list) -> None:
    if SHEET_DETAIL in wb.sheetnames:
        del wb[SHEET_DETAIL]
    ws = wb.create_sheet(SHEET_DETAIL)

    headers = ["Object Gold", "Profile", "Mã", "Nhóm", "Tiêu chí", "Kết quả máy",
               "Ghi vào Excel", "Bằng chứng / số dòng", "Cách sửa", "File"]
    widths = [34, 17, 7, 24, 44, 12, 13, 82, 88, 38]
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = HEAD
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[c.column_letter].width = w
    ws.freeze_panes = "A2"

    colors = {FAIL: NAVY, WARN: BLUE, PASS: GREEN}
    r = 2
    for ctx, findings in results:
        for rid in sorted(findings, key=_sort_key):
            if not in_checklist(rid):     # nhom X khong nam trong checklist OCB
                continue
            f = findings[rid]
            rule = RULES[rid]
            vals = [ctx.label, ctx.profile, rid, rule.group,
                    rule_text.title(rid, rule.title), f.status,
                    _cell_value(f.status),
                    f.note(), rule_text.fix(rid), ctx.path]
            for i, v in enumerate(vals, start=1):
                c = ws.cell(row=r, column=i, value=v)
                c.font = ARIAL
                c.alignment = Alignment(vertical="top", wrap_text=(i in (8, 9)))
            sc = ws.cell(row=r, column=6)
            if f.status in colors and f.status != NA:
                sc.font = Font(name="Arial", size=11, bold=True, color=colors[f.status])
            r += 1


def _sort_key(rid: str):
    grp, num = rid.split(".")
    return (grp, int(num))
