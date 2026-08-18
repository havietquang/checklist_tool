"""Doc workbook mapping thiet ke (*_Silver_to_Gold.xlsx / *_mapping*.xlsx) lam INPUT thu 2.

Cau truc workbook duoc ho tro (theo mau dang dung trong 8_VIE):

  Sheet chinh (ten = ten object Gold)
    Dong 1  : '<OBJ>' + mo ta, kem nhan loai '[ BulkLoad ]' (bang) hoac '[ VIEW ]'
    Block   : 'JOIN SCHEMA ...'   header: # | TABLE / VIEW | ALIAS | JOIN TYPE | ON / CONDITIONS | NOTE
    Block   : 'FIELD MAPPING ...' header: STT | FIELD | DATA TYPE | TYPE | TRANSFORM | TABLE / VIEW SOURCE

  Sheet 'SQL' / 'Script SQL' / 'Code'   <- chi dump code, khong doc lam dac ta
"""
from __future__ import annotations

import os
import re
import unicodedata
from dataclasses import dataclass, field

import openpyxl


def unaccent(text: str) -> str:
    """Bo dau tieng Viet de so chuoi khong phu thuoc dau ('Code mới' -> 'Code moi')."""
    stripped = "".join(c for c in unicodedata.normalize("NFD", text or "")
                       if not unicodedata.combining(c))
    return stripped.replace("đ", "d").replace("Đ", "D")

JOIN_HDR_RE = re.compile(r"TABLE\s*/\s*VIEW", re.I)
FIELD_HDR_RE = re.compile(r"^\s*(STT|#)\s*$", re.I)
# Sheet chi chua dump code, khong phai dac ta mapping -> bo qua khi doc.
# Chap nhan ca gach duoi / gach ngang lam dau phan cach ('SCRIPT_SQL', 'Script-SQL'):
# \b KHONG khop sau 'SCRIPT' trong 'SCRIPT_SQL' vi '_' cung la ky tu tu, truoc day lam
# ca workbook bi bo qua vi tool khong nhan ra do la sheet chua code.
SQL_SHEET_RE = re.compile(r"^\s*(script|sql|code)([\s_\-.]|$)", re.I)
NOISE_RE = re.compile(r"(_Silver_to_Gold|_mapping(_code)?|_VIEW_mapping|_backup|_new|_v\d+|"
                      r"_\d{8}|\s*\([^)]*\)|\.fix|\.backup)", re.I)


@dataclass
class SourceTable:
    name: str
    alias: str = ""
    join_type: str = ""
    condition: str = ""
    note: str = ""


@dataclass
class FieldMap:
    field: str
    data_type: str = ""
    transform_type: str = ""     # 1:1 / Transform / Lookup / Aggregate / Constant / Filter
    transform: str = ""
    source: str = ""


@dataclass
class MappingDoc:
    path: str
    object_name: str = ""
    object_kind: str = ""                       # BULKLOAD | VIEW | ''
    sources: list = field(default_factory=list)  # list[SourceTable]
    fields: list = field(default_factory=list)   # list[FieldMap]
    sql_sheets: list = field(default_factory=list)
    notes: list = field(default_factory=list)

    @property
    def field_names(self) -> list:
        """Danh sach cot dich DA CHUAN HOA de doi chieu voi cot dau ra cua SQL:

        1. Bo dau nhay / backtick nguoi lam mapping go kem quanh ten cot dac biet
           (`CN/PGD` / 'CN/PGD' / CN/PGD' -> CN/PGD). sqlglot doc SQL cung tra ve ten
           tran khong con dau nhay, khong bo thi so lech oan.
        2. Bo TRUNG LAP: khi SQL co UNION ALL nhieu nhanh, thiet ke thuong khai mot block
           FIELD MAPPING cho TUNG nhanh -> cung mot tap cot lap lai nhieu lan (vd
           V_P_REGION_CPHD_UPL: 7 cot khai 2 lan = 14 dong). Object chi co 7 cot that,
           dem ca 14 dong se bao lech mapping oan."""
        out, seen = [], set()
        for f in self.fields:
            name = re.sub(r"[`'\"]", "", f.field or "").strip().upper()
            if name and name not in seen:
                seen.add(name)
                out.append(name)
        return out

    @property
    def source_names(self) -> list:
        return [s.name.upper() for s in self.sources]


EXTS = (".xlsx", ".xlsm", ".xls", ".sql")


def stt_from_filename(path: str):
    """So thu tu dau ten file workbook thiet ke, vd
      '067. OCB_GOLD_TCKH_AR_DIM_HOP.xlsx' -> 67
    Dung de sap xep + dien cot STT cua checklist theo DUNG thu tu workbook thiet ke da
    duoc OCB danh so tu truoc, thay vi danh lai tu 1 theo thu tu tool doc file (gay lech
    voi STT ma nguoi review dang doi chieu tren workbook goc). Tra ve None neu ten file
    khong co so thu tu dau (vd file .sql thuong, khong theo quy uoc OCB_GOLD_...)."""
    stem = os.path.splitext(os.path.basename(path))[0]
    m = re.match(r"\s*(\d+)[.)\-]?\s", stem)
    return int(m.group(1)) if m else None


def object_from_filename(path: str) -> str:
    """Tach ten object THAT (object 'chinh chu') tu ten file workbook.

    Quy uoc dat ten OCB: '[<stt>. ]OCB_GOLD_<MODULE>_<TEN_OBJECT>_<PIC>.xlsx', vi du
      '079. OCB_GOLD_TCKH_CB_OU_DIM_HOP.xlsx'      -> 'CB_OU_DIM'
      '088. OCB_GOLD_TTDL_TBL_BPM_QLNS_LINH.xlsx'  -> 'TBL_BPM_QLNS'
      'OCB_GOLD__TCKH__V_SLNV_UPL__HOP.xlsx'       -> 'V_SLNV_UPL'   (ban cu dung '__')
    Bo: so thu tu dau, tien to 'OCB_GOLD_', doan <MODULE> dau tien va <PIC> cuoi cung -
    ca hai KHONG phai ten object. Dung cho 2 viec: (1) doan target khi SQL khong co
    INSERT/CREATE (core._target_of), (2) xac dinh workbook CHINH CHU cua 1 object khi
    cung 1 object bi khai SQL o nhieu workbook (_pick_script)."""
    parts = _filename_parts(path)
    return "_".join(parts[1:-1]) if len(parts) >= 3 else "_".join(parts)


def _filename_parts(path: str) -> list:
    """Ten file workbook -> [<MODULE>, <phan ten object>..., <PIC>] da bo rac."""
    stem = os.path.splitext(os.path.basename(path))[0]
    stem = re.sub(r"^\s*\d+[.)\-]?\s*", "", stem)        # bo so thu tu '079. '
    stem = re.sub(r"_{2,}", "_", stem).strip(" _-")       # ban cu dung '__'
    stem = re.sub(r"^OCB_GOLD_", "", stem, flags=re.I)
    return stem.split("_")


def owns(path: str, obj: str) -> bool:
    """Workbook `path` co phai la workbook CHINH CHU cua object `obj` khong.

    So phan con lai cua ten file sau khi bo so thu tu / 'OCB_GOLD_' / <MODULE>, chap nhan
    ca truong hop ten file THIEU hau to PIC:
      '027. OCB_GOLD_TCKH_V_SBV_BAOANH.xlsx' -> 'V_SBV_BAOANH' so huu V_SBV (khong so huu
      V_SBV_1); '058. OCB_GOLD_TCKH_V_SBV_1.xlsx' -> 'V_SBV_1' so huu dung V_SBV_1.
    Chinh xac hon object_from_filename() vi khong phai doan doan nao la PIC."""
    rest = "_".join(_filename_parts(path)[1:]).upper()
    name = clean_key(obj)
    return bool(name) and (rest == name or rest.startswith(name + "_"))


def pic_from_filename(path: str) -> str:
    """PIC (nguoi phu trach) = doan CUOI cua ten file workbook, vd
      '079. OCB_GOLD_TCKH_CB_OU_DIM_HOP.xlsx'                  -> 'HOP'
      '062. OCB_GOLD_TCKH_WORKING_DAY_BAOANH.xlsx'             -> 'BAOANH'
      '013. OCB_GOLD_TCKH_V_LOAN_BY_PRD_DAILY_PHAT Copy.xlsx'  -> 'PHAT'
    Dung de dien cot 'PIC review' cua checklist theo tung object, thay vi mot gia tri
    chung cho ca batch (moi object do mot nguoi lam). Tra ve '' neu ten file khong theo
    quy uoc (khong du 3 doan) de khong dien bua vao file gui OCB."""
    parts = _filename_parts(path)
    if len(parts) < 3:
        return ""
    pic = re.sub(r"\s*(copy|\(\d+\)|bak|backup)\s*$", "", parts[-1], flags=re.I)
    return pic.strip(" _-")


def clean_key(name: str) -> str:
    """Chuan hoa NHE cho ten object DA DOC THANG tu noi dung (ws.title, block header,
    ten cot Object trong sheet Script SQL, ten bang trong CREATE/INSERT...): chi
    uppercase + bo khoang trang/gach noi thua, KHONG bo noise versioning nhu NOISE_RE.
    Ly do: ten object o day co the THAT SU ket thuc bang '_NEW'/'_V2' nhu 1 phan ten
    nghiep vu that (vd V_CUST_STATUS_NEW la object KHAC HAN V_CUST_STATUS, khong phai
    ban '_new' cua no) - dung norm_key() se lam mat hau to that, gay dam khoa voi 1
    object khac. Dung norm_key() day du chi cho TEN FILE (noi noise versioning that su
    la rac, vd 'GOOD_FCT_Silver_to_Gold.xlsx').

    VAN bo tien to schema kieu 'dbo.V_SLGD' -> 'V_SLGD' (lay doan cuoi sau dau '.') -
    khac voi noise versioning, tien to schema la quy uoc SQL ro rang (schema.table),
    khong phai doan ten nghiep vu that nhu '_NEW'.

    Cung bo phan chu thich trong ngoac ma nguoi lam mapping ghi kem, vd o Table/View ghi
    'v_cdtk_daily_1 (GOLD)' -> 'V_CDTK_DAILY_1' (ten bang SQL khong bao gio co dau ngoac,
    nen day chac chan la chu thich; giu lai se khong khop duoc voi workbook chinh chu)."""
    stem = re.sub(r"\([^)]*\)", " ", name or "").strip(" _-")
    return stem.split(".")[-1].strip(" _-").upper()


def norm_key(name: str) -> str:
    """Chuan hoa ten file/object de ghep mapping voi file SQL.
    'OCBRT.GOOD_FCT' / 'GOOD_FCT (Gold)' / 'GOOD_FCT_Silver_to_Gold.xlsx' -> 'GOOD_FCT'."""
    stem = os.path.basename(name)
    for ext in EXTS:                       # chi bo duoi file that su, khong dung splitext
        if stem.lower().endswith(ext):     # ('1.0 V_CDR_AMT' bi splitext cat thanh '1')
            stem = stem[: -len(ext)]
            break
    stem = NOISE_RE.sub("", stem).strip(" _-")
    return stem.split(".")[-1].strip(" _-").upper()


def discover(dirs: list) -> dict:
    """Quet thu muc, tra ve {khoa_object: MappingDoc}.

    1 workbook co the mo ta NHIEU object (moi sheet 1 object, hoac 2 block canh nhau
    tren cung 1 sheet) -> index theo ten object, khong theo ten file.
    """
    import glob
    out = {}
    for d in dirs:
        for p in sorted(glob.glob(os.path.join(d, "*.xlsx"))):
            base = os.path.basename(p)
            if base.startswith("~$") or "Checklist_Review" in base:
                continue
            try:
                docs = read_all(p)
            except Exception as exc:            # noqa: BLE001
                doc = MappingDoc(path=p, object_name=norm_key(p))
                doc.notes.append(f"khong doc duoc workbook: {type(exc).__name__}: {exc}")
                docs = [doc]
            for doc in docs:
                key = clean_key(doc.object_name)
                old = out.get(key)
                if old is None or _better(doc, old):
                    out[key] = doc
    return out


def _better(new: MappingDoc, old: MappingDoc) -> bool:
    """Uu tien: (1) ten FILE khop ten object, (2) co field mapping, (3) nhieu field hon,
    (4) file moi hon.
    (1) quan trong nhat: RT_PL_DTL phai lay tu RT_PL_DTL_Silver_to_Gold.xlsx, khong lay tu
    sheet '1.10a RT_PL_DTL' nam trong workbook cua object khac.

    Dung object_from_filename() (bo dung tien to 'OCB_GOLD_<MODULE>_' + hau to '_<PIC>')
    de so khop, KHONG dung norm_key(path) truc tiep - norm_key chi bo cac noise ver-hoa
    ('_v2', '_backup'...) chu khong bo duoc tien to/hau to theo quy uoc dat ten OCB, nen
    voi quy uoc nay norm_key(path) gan nhu KHONG BAO GIO khop norm_key(object_name) cho
    BAT KY file nao -> tieu chi (1) bi vo hieu hoan toan, luon roi xuong so mtime (4) mot
    cach ngau nhien giua cac workbook co CUNG 1 block object trung lap (vd nhieu workbook
    deu co copy-dan block 'V_FTP_RATE' de tham khao) - day la nguyen nhan gay nham workbook
    chinh chu nhieu lan trong thuc te."""
    def key(d: MappingDoc):
        fname_obj = object_from_filename(d.path).strip(" _-").upper()
        return (fname_obj == norm_key(d.object_name).upper(),
                len(d.fields) > 0, len(d.fields))

    if key(new) != key(old):
        return key(new) > key(old)
    return os.path.getmtime(new.path) > os.path.getmtime(old.path)


def read(path: str) -> MappingDoc:
    """Tra ve doc dau tien (tuong thich nguoc; dung read_all khi workbook co nhieu object)."""
    docs = read_all(path)
    return docs[0] if docs else MappingDoc(path=path, object_name=norm_key(path))


def read_all(path: str) -> list:
    wb = openpyxl.load_workbook(path, data_only=True)
    docs, sql_sheets = [], []
    # Workbook dang lineage (Summary + cac sheet '1.x <bang>') mo ta MOT object:
    # sheet '1.0 <OBJ>' la object dich, cac sheet 1.x con lai la bang nguon -> gop 1 doc.
    lineage = MappingDoc(path=path, object_name="")

    for ws in wb.worksheets:
        if SQL_SHEET_RE.match(ws.title.strip()):
            sql_sheets.append(ws.title)
        elif _read_lineage_sheet(ws, _rows(ws), lineage):
            continue
        else:
            docs += _read_sheet(ws, path)

    if lineage.fields or lineage.sources:
        if not lineage.object_name:
            lineage.object_name = norm_key(path)
        docs.insert(0, lineage)

    if not docs:                               # workbook chi co sheet SQL
        docs = [MappingDoc(path=path, object_name=norm_key(path))]
    for doc in docs:
        doc.sql_sheets = sql_sheets
    return docs


LINEAGE_HDR = ("target column", "transform source table")
LV0_SHEET_RE = re.compile(r"^1\.0\s+(.+)$")


BLOCK_RE = re.compile(r"(JOIN\s+SCHEMA|FIELD\s+MAPPING)\s*[-—:]*\s*(.*)$", re.I)
# ten object hop le: dinh danh bang/view, khong khoang trang (loai 'cac bang nguon')
OBJ_NAME_RE = re.compile(r"^[A-Za-z_][A-Za-z0-9_.]{2,}$")
# nhan TANG thiet ke (LV0/LV1/LV2...), khong phai ten object that - vd tieu de
# 'JOIN SCHEMA — LV1' / 'FIELD MAPPING — LV1 (V_BRANCH_LIST)'. Nhieu workbook khac nhau
# dung chung nhan nay -> neu nhan lam ten object se dam vao nhau thanh 1 khoa "LV1" duy
# nhat, cac object khac bi de mat. Khi gap nhan nay phai bo qua de _read_sheet roi lai
# tu ws.title (ten sheet, luon la ten object that trong kieu workbook nay).
LEVEL_LABEL_RE = re.compile(r"^LV\s*\d+$", re.I)


def _rows(ws) -> list:
    return [[("" if c is None else str(c).strip()) for c in r]
            for r in ws.iter_rows(values_only=True)]


def _read_sheet(ws, path: str) -> list:
    """1 sheet -> 1 hoac nhieu MappingDoc (2 block canh nhau theo cot = 2 object)."""
    rows = _rows(ws)
    if not rows:
        return []

    # tim cac cot bat dau block; nhieu cot khac nhau = nhieu object canh nhau
    starts, names = set(), {}
    for r in rows:
        for ci, cell in enumerate(r):
            m = BLOCK_RE.match(cell or "")
            if m:
                starts.add(ci)
                obj = re.split(r"\s*\(|\s{2,}", m.group(2).strip())[0].strip(" —-:")
                # chi nhan neu la dinh danh bang/view thuc su; loai mo ta kieu
                # 'JOIN SCHEMA — cac bang nguon' VA loai nhan tang LV0/LV1/LV2 (khong
                # phai ten object - de roi ben duoi tu dong dung ten sheet).
                if ci not in names and OBJ_NAME_RE.match(obj) and not LEVEL_LABEL_RE.match(obj):
                    names[ci] = obj
    if not starts:
        return []

    bounds = sorted(starts)
    docs = []
    for i, c0 in enumerate(bounds):
        c1 = bounds[i + 1] if i + 1 < len(bounds) else max(len(r) for r in rows)
        # Sheet chi co 1 block (khong tach nhieu object canh nhau): theo dung quy uoc
        # cua kieu workbook nay (xem docstring dau file) TEN SHEET moi la ten object
        # chuan - dang tin hon chu trong tieu de block vi tieu de hay bi go tat/thieu
        # hau to (vd sheet 'V_SLNV_UPL' nhung tieu de ghi nham 'FIELD MAPPING — V_SLNV',
        # thieu '_UPL'). Chi khi sheet co NHIEU block canh nhau (nhieu object dung chung
        # 1 sheet) moi bat buoc phai doc ten tu tieu de block de phan biet tung object.
        name = ws.title.strip() if len(bounds) == 1 else (names.get(c0) or ws.title.strip())
        doc = MappingDoc(path=path, object_name=name)
        _fill_from_region(doc, [r[c0:c1] for r in rows], " ".join(rows[0]))
        if doc.fields or doc.sources:
            docs.append(doc)
    return docs


def _fill_from_region(doc: MappingDoc, rows: list, sheet_head: str) -> None:
    head = (sheet_head + " " + " ".join(rows[0]))[:300]
    if re.search(r"\[\s*VIEW\s*\]|\bVIEW\b", head, re.I):
        doc.object_kind = "VIEW"
    elif re.search(r"\[\s*BulkLoad\s*\]|\bTABLE\b", head, re.I):
        doc.object_kind = "BULKLOAD"

    mode, cols = None, {}
    for r in rows:
        joined = " ".join(r)
        if not joined.strip():
            continue
        if re.search(r"JOIN\s+SCHEMA", joined, re.I):
            mode, cols = "await_join", {}
            continue
        if re.search(r"FIELD\s+MAPPING", joined, re.I):
            mode, cols = "await_field", {}
            continue
        # Header nhan theo TEN cot, khong bat buoc phai co cot STT
        if mode == "await_join" and JOIN_HDR_RE.search(joined):
            resolved = _resolve(r, JOIN_COLS)
            if "name" in resolved:
                mode, cols = "join", resolved
                continue
        if mode == "await_field":
            resolved = _resolve(r, FIELD_COLS)
            if "field" in resolved:
                mode, cols = "field", resolved
                continue

        if mode == "join":
            name = _at(r, cols, "name")
            if name and not JOIN_HDR_RE.search(name):
                doc.sources.append(SourceTable(
                    name=name, alias=_at(r, cols, "alias"),
                    join_type=_at(r, cols, "join_type"), condition=_at(r, cols, "condition"),
                    note=_at(r, cols, "note")))
        elif mode == "field":
            name = _at(r, cols, "field")
            if name and not re.match(r"^(field|target\s+column)", name, re.I):
                doc.fields.append(FieldMap(
                    field=name, data_type=_at(r, cols, "data_type"),
                    transform_type=_at(r, cols, "transform_type"),
                    transform=_at(r, cols, "transform"), source=_at(r, cols, "source")))


def _read_lineage_sheet(ws, rows: list, doc: MappingDoc) -> bool:
    """Format lineage: sheet '1.x <bang>' voi header
    STT | Target table | Target column | Alias | Type | Transform | Transform source table | ...
    Sheet '1.0 <OBJ>' la object dich (LV0) -> lay field mapping; cac sheet 1.x con lai la nguon."""
    hdr_idx = next((i for i, r in enumerate(rows[:4])
                    if all(any(h.lower() == c for c in (x.lower() for x in r)) for h in LINEAGE_HDR)), None)
    if hdr_idx is None:
        return False

    header = [h.lower() for h in rows[hdr_idx]]
    col = {name: header.index(name) for name in
           ("target column", "type", "transform", "transform source table", "condition join")
           if name in header}
    lv0 = LV0_SHEET_RE.match(ws.title.strip())

    if lv0:
        doc.object_name = lv0.group(1).strip()
        if not doc.object_kind:
            doc.object_kind = "VIEW" if lv0.group(1).upper().startswith("V_") else "BULKLOAD"
        for r in rows[hdr_idx + 1:]:
            name = _cell(r, col.get("target column", -1))
            if not name or name.upper() in ("TARGET COLUMN",):
                continue
            doc.fields.append(FieldMap(
                field=name,
                transform_type=_cell(r, col.get("type", -1)),
                transform=_cell(r, col.get("transform", -1)),
                source=_cell(r, col.get("transform source table", -1))))
    else:
        src = re.sub(r"^1\.\d+[a-z]?\s+", "", ws.title.strip())
        cond = next((_cell(r, col["condition join"]) for r in rows[hdr_idx + 1:]
                     if "condition join" in col and _cell(r, col["condition join"])), "")
        doc.sources.append(SourceTable(name=src, condition=cond, note="tu sheet lineage"))
    return True


# ---------------------------------------------------------------- doc SQL tu sheet Script
# Cho phep hau to nhan TANG sau ten cot ('View/Table LV1', 'Table/View Lv2') VA tien to
# 'Ten '/'Tên ' truoc ten cot ('Tên View/Table'): nhieu workbook ghi kem tang thiet ke o
# tieu de cot, hoac them chu 'Tên' cho de doc. Khong nhan duoc thi _scan_script_sheet tra
# ve rong -> CA workbook bi bo qua IM LANG (da tung lam mat 5 file BPM_* + 4 file
# V_TB_AR_DTL* khoi luot cham vi header ghi 'Tên View/Table' thay vi 'View/Table').
OBJ_HDR_RE = re.compile(r"^(t[êe]n\s+)?(object|view\s*/\s*table|table\s*/\s*view|view|table)"
                        r"(\s*lv\s*\d+)?$", re.I)
SCRIPT_HDR_RE = re.compile(r"script|sql|code", re.I)
TYPE_HDR_RE = re.compile(r"^(type|loai)$", re.I)
# Chi lay code MOI (Databricks). 'Code cu' la T-SQL on-prem, khong phai thu can cham.
# So sau khi BO DAU vi cot Type viet tieng Viet co dau ('Code moi' / 'Code cu').
NEW_CODE_RE = re.compile(r"moi|new|gold|dbx|databricks", re.I)
SQL_BODY_RE = re.compile(r"\b(SELECT|CREATE|INSERT|MERGE|WITH|DELETE)\b", re.I)
# Dau hieu code CU on-prem (T-SQL SQL Server). Dung khi o cot Type BO TRONG: luc do
# khong the loc theo nhan 'Code moi'/'Code cu' nua, phai nhin chinh SQL. Co workbook de
# trong ca cot Type (vd '010. V_FTP_003') -> truoc day bi loc sach, mat ca file.
OLD_CODE_RE = re.compile(r"\[dbo\]|\bdbo\.", re.I)

# O Script SQL duoc phep ghi DUONG DAN toi 1 file .sql THAT thay vi dan nguyen code vao
# nhieu o (dep hon, va luon dong bo voi code that trong repo, khong so lech ban). Nhan
# dien: ca o (sau khi noi cac dong cung khoi) chi la 1 DONG DUY NHAT, ket thuc bang .sql.
SQL_FILE_REF_RE = re.compile(r'^["\']?(?P<path>[^\r\n"\']+\.sql)["\']?$', re.I)
# Goc repo, dung de quy doi duong dan tuong doi ghi trong o ve file that tren dia.
_ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))


def _resolve_sql_ref(ref: str, base_dir: str) -> str | None:
    """Quy doi duong dan ghi trong o Script SQL ve file that tren dia.

    Thu lan luot: duong dan tuyet doi -> tuong doi voi THU MUC CHUA WORKBOOK (thuong go
    duong dan ngan tinh tu day) -> tuong doi voi GOC REPO. None neu khong tim thay dau."""
    ref = ref.strip()
    candidates = [ref] if os.path.isabs(ref) else [
        os.path.join(base_dir, ref), os.path.join(_ROOT_DIR, ref)]
    return next((c for c in candidates if os.path.isfile(c)), None)


# Workbook bi bo qua khi quet sheet Script (ly do). run_check doc de in canh bao -
# im lang bo file la nguy hiem nhat: object bien mat khoi ket qua ma khong ai biet.
SKIPPED_FILES: list = []
# Cot Script SQL ghi duong dan file .sql nhung KHONG tim thay file do tren dia (go sai
# ten/duong dan). Rieng voi SKIPPED_FILES vi day la loi 1 object, khong phai ca workbook.
SQL_REF_ISSUES: list = []
# O Script SQL bi bo qua vi khong chua than SQL nao va cung khong phai phan tiep cua object
# nao. Thuong la o ghi chu/mo ta - nhung cung co the la CODE BI CAT DOAN ma go sai cach.
DROPPED_CELLS: list = []


COPY_FILE_RE = re.compile(r"\b(copy|bak|backup|old|cu)\b|\bcopy\b|\(\d+\)", re.I)


def discover_scripts(dirs: list, conflicts: list | None = None,
                     variants: dict | None = None) -> dict:
    """Quet workbook thiet ke, tra ve {khoa_object: (duong_dan_workbook, sheet, sql)}.

    Dung khi chay --from-excel: SQL lay thang tu sheet 'Script SQL' cua workbook,
    khong can thu muc input/sql. Chi lay dong co Type la 'Code moi' (code Databricks).

    Bang DUNG CHUNG (CB_OU_DIM, BRANCH_LIST, HOLIDAY...) thuong bi dan lai SQL vao sheet
    Script cua NHIEU workbook tieu thu. Truoc day lay ban gap dau tien theo alphabet nen
    ban trong workbook CHINH CHU co the bi bo qua im lang; nay _pick_script() luon uu tien
    ban trong workbook mang dung ten object. Truyen 1 list vao `conflicts` de nhan lai
    danh sach object bi khai o nhieu noi (kem co `differs` khi cac ban LECH NOI DUNG).

    Truyen 1 dict vao `variants` de nhan lai TAT CA cac ban tim duoc
    ({khoa_object: [(workbook, sheet, sql), ...]}), khong chi ban duoc chon - sync.py can
    day du de chi ro noi nao dang giu ban cu."""
    import glob
    SKIPPED_FILES.clear()
    SQL_REF_ISSUES.clear()
    DROPPED_CELLS.clear()
    found: dict[str, list] = {}
    for d in dirs:
        for p in sorted(glob.glob(os.path.join(d, "*.xlsx"))):
            base = os.path.basename(p)
            if base.startswith("~$") or "Checklist_Review" in base:
                continue
            try:
                wb = openpyxl.load_workbook(p, data_only=True)
            except Exception as exc:                              # noqa: BLE001
                SKIPPED_FILES.append((base, f"khong mo duoc workbook: {type(exc).__name__}"))
                continue
            sheets = [ws for ws in wb.worksheets if SQL_SHEET_RE.match(ws.title.strip())]
            if not sheets:
                SKIPPED_FILES.append((base, "khong co sheet Script/SQL/Code nao"))
                continue
            got = 0
            for ws in sheets:
                for key, sql in _scan_script_sheet(ws, os.path.dirname(p), base).items():
                    if key:
                        found.setdefault(key, []).append((p, ws.title, sql))
                        got += 1
            # Workbook CO sheet Script nhung khong lay duoc dong nao: truoc day bi bo qua
            # IM LANG -> ca file bien mat khoi luot cham ma khong ai biet. Phai bao ro.
            if not got:
                SKIPPED_FILES.append((
                    base, f"sheet {[ws.title for ws in sheets]} khong doc duoc dong code nao"
                          " (thieu cot Type / View-Table / Script SQL, hoac khong co dong"
                          " Type='Code moi')"))

    if variants is not None:
        variants.clear()
        variants.update(found)
    out = {}
    for key, cands in found.items():
        out[key] = _pick_script(key, cands)
        if conflicts is not None and len(cands) > 1:
            conflicts.append({
                "object": key,
                "chosen": out[key][0],
                "others": [p for p, _, _ in cands if p != out[key][0]],
                "differs": len({sql for _, _, sql in cands}) > 1,
                "owned": owns(out[key][0], key),
            })
    return out


def _pick_script(key: str, cands: list) -> tuple:
    """Chon ban SQL nao khi 1 object duoc khai o nhieu workbook.

    Thu tu uu tien: (1) workbook CHINH CHU - ten file mang dung ten object, vd
    '079. OCB_GOLD_TCKH_CB_OU_DIM_HOP.xlsx' la chinh chu cua CB_OU_DIM; (2) khong phai
    file ban sao ('... Copy.xlsx', '(1)', '_bak'); (3) alphabet cho on dinh ket qua."""
    owned = [c for c in cands if owns(c[0], key)]
    pool = owned or cands
    fresh = [c for c in pool if not COPY_FILE_RE.search(os.path.basename(c[0]))]
    pool = fresh or pool
    return sorted(pool, key=lambda c: os.path.basename(c[0]))[0]


HDR_SCAN_ROWS = 6            # do sau toi da di tim dong tieu de


def _script_header(rows: list) -> tuple:
    """Tim DONG TIEU DE cua sheet Script, tra ve (chi_so_dong, i_type, i_obj, i_sql).

    Khong the mac dinh tieu de o dong 1: nhieu workbook chua 1-2 dong trong (hoac dong
    tieu de phu) o tren truoc khi vao bang - vd '006. V_CUSTOMER_SEGMENT_CMB' co dong 1
    rong, tieu de o dong 2. Truoc day doc cung rows[0] nen gap dong rong la bo qua CA
    workbook trong im lang. Quet vai dong dau, lay dong dau tien co du cot Object + Script."""
    for idx, header in enumerate(rows[:HDR_SCAN_ROWS]):
        i_type = next((i for i, h in enumerate(header) if TYPE_HDR_RE.match(h)), None)
        i_obj = next((i for i, h in enumerate(header) if OBJ_HDR_RE.match(h)), None)
        i_sql = next((i for i, h in enumerate(header)
                      if SCRIPT_HDR_RE.search(h) and i not in (i_type, i_obj)), None)
        if i_obj is not None and i_sql is not None:
            return idx, i_type, i_obj, i_sql
    return -1, None, None, None


def _scan_script_sheet(ws, base_dir: str = "", label: str = "") -> dict:
    """{khoa_object: sql} lay tu sheet dang bang Type | Object | Script SQL.

    Ho tro 2 cach ghi code thay vi dan het vao 1 o (o Excel toi da ~32.767 ky tu):
    (1) CAT THANH NHIEU O/NHIEU DONG - chap nhan CA HAI kieu ghi, deu duoc NOI VAO CUOI
        (xuong dong) theo dung thu tu xuat hien trong sheet:
          - dong dau co Object (+ Type='Code moi'), cac dong TIEP THEO DE TRONG Object;
          - hoac GHI LAI ten Object o moi dong (vd '077. LOAN_SUMMARY_LIST' cat 3 o
            32701 + 22322 + 11241 ky tu, dong nao cung ghi ten).
    (2) GHI DUONG DAN toi 1 file .sql THAT: o Script SQL chi ghi 1 dong duy nhat la
        duong dan file (vd 'zonec_datamart/.../pst_entr_fct.sql') - tool tu doc noi dung
        file do, luon dong bo voi code that trong repo thay vi phai dan lai. `base_dir`
        la thu muc chua workbook, dung de quy doi duong dan tuong doi; `label` chi de
        ghi vao SQL_REF_ISSUES khi khong tim thay file."""
    rows = _rows(ws)
    if not rows:
        return {}
    hdr_idx, i_type, i_obj, i_sql = _script_header(rows)
    if hdr_idx < 0:
        return {}

    out: dict[str, str] = {}
    cur_key, cur_typ, cur_lines = None, "", []

    def flush() -> None:
        # Chi den khi HET khoi (gap Object moi hoac het bang) moi ghep cac dong lai va
        # kiem tra - dong dau 1 khoi thuong chi la 'USE CATALOG...'/'USE SCHEMA...', chua
        # co CREATE/SELECT nen KHONG THE xet SQL_BODY_RE tung dong rieng le duoc.
        nonlocal cur_key, cur_typ, cur_lines
        if cur_key and cur_lines:
            joined = "\n".join(cur_lines)
            ref = SQL_FILE_REF_RE.match(joined.strip()) if len(cur_lines) == 1 else None
            if ref:
                path = _resolve_sql_ref(ref.group("path"), base_dir)
                if path:
                    with open(path, encoding="utf-8", errors="replace") as fh:
                        joined = fh.read()
                else:
                    SQL_REF_ISSUES.append(
                        (label, f"{cur_key}: khong tim thay file '{ref.group('path')}'"
                                f" (da thu tuong doi voi thu muc workbook va goc repo)"))
                    joined = ""
            # Khoi tiep cua MOT object da doc duoc code (nguoi lam khai lai ten object o
            # dong sau thay vi de trong cot View/Table) -> nhan luon, KHONG doi phai co
            # SELECT/CREATE. Doan giua cau SQL nhu 'FROM ... WHERE ...' hay 'GROUP BY ...'
            # khong he chua tu khoa nao trong SQL_BODY_RE: truoc day bi bo IM LANG, code
            # trong o do mat sach ma khong ai biet (test_split: doc 3/7 dong).
            cont = cur_key in out
            if joined and (cont or SQL_BODY_RE.search(joined)):
                skip = (not NEW_CODE_RE.search(cur_typ)) if cur_typ else bool(OLD_CODE_RE.search(joined))
                if not skip:
                    out[cur_key] = out[cur_key] + "\n" + joined if cont else joined
            elif joined:
                # Khong phai phan tiep, cung khong co than SQL nao -> that su bo qua o nay.
                # Phai bao ro: neu day dung la code bi cat doan thi nguoi lam moi biet ma
                # sua cach ghi (de trong cot View/Table o cac dong sau).
                DROPPED_CELLS.append(
                    (label, f"{cur_key}: bo qua 1 o Script SQL khong chua SELECT/CREATE/"
                            f"INSERT/MERGE/WITH/DELETE ({len(joined)} ky tu, bat dau"
                            f" '{joined.strip()[:40]}...')"))
        cur_key, cur_typ, cur_lines = None, "", []

    for r in rows[hdr_idx + 1:]:
        sql = _unquote_cell(r[i_sql] if i_sql < len(r) else "")
        raw_obj = _cell(r, i_obj) if i_obj is not None else ""
        key = clean_key(raw_obj) if raw_obj else None

        if key is None:
            # Khong khai Object -> phan tiep cua khoi dang mo (o truoc da het cho).
            if sql and cur_key:
                cur_lines.append(sql)
            continue

        # Co khai Object -> ket thuc khoi truoc, mo 1 khoi moi.
        flush()
        cur_key = key
        cur_typ = unaccent(_cell(r, i_type)) if i_type is not None else ""
        if sql:
            cur_lines.append(sql)
    flush()
    return out


def _unquote_cell(sql) -> str:
    """Bo cap dau nhay kep BAO NGOAI o Script SQL (kieu quoting cua CSV).

    Code copy tu file .csv/.txt dan vao Excel hay bi boc ca cap '"' o dau va cuoi O. Voi
    code dan 1 o thi vo hai, nhung code dai phai cat NHIEU O: noi lai se thanh
    '... sched_day ), " " avi as ( ...' - dau nhay lot vao GIUA cau SQL. Hau qua:
      - sqlglot parse ra cay khac (hoac loi cu phap) -> cham diem sai;
      - --sync-report bao LECH voi ban chuan chi vi 4 ky tu nay (dung case
        LOAN_SUMMARY_LIST: 034/035/046 vs 077 - noi dung y het, chi lech cap '"').
    Chi bo khi o BAT DAU va KET THUC bang '"' - do la dau hieu quoting, khong phai SQL."""
    if not isinstance(sql, str):
        return sql or ""
    s = sql.strip()
    if len(s) > 2 and s.startswith('"') and s.endswith('"'):
        return s[1:-1]
    # cat nhieu o: o dau chi co '"' mo, o cuoi chi co '"' dong -> bo not ve dung 1 phia
    if len(s) > 1 and s.startswith('"') and '"' not in s[1:]:
        return s[1:]
    if len(s) > 1 and s.endswith('"') and '"' not in s[:-1]:
        return s[:-1]
    return sql


def _cell(row: list, idx: int) -> str:
    return row[idx].replace("\n", " ").strip() if 0 <= idx < len(row) and row[idx] else ""


# Header co nhieu bien the giua cac workbook -> nhan cot theo TEN, khong theo vi tri.
# Moi khoa: danh sach regex thu tu uu tien.
FIELD_COLS = {
    # Cac pattern chat truoc uu tien tranh nham voi cot khac; pattern long r"field"
    # o cuoi de bat luon bien the co chen them nhan tang nhu 'Field Lv1\n(Output Col)'
    # ('Lv1' chen giua lam cac pattern chat o tren khong khop) - co chu 'field' o dau
    # la du, khong can dung dinh dang.
    "field": [r"^fields?$", r"^field\s*\(output", r"^target\s+column$", r"^column$",
              r"\bfield\b"],
    "data_type": [r"^data\s*type$", r"^kieu\s*du\s*lieu$"],
    "transform_type": [r"^transform\s*type$", r"^type$", r"^transform$", r"^loai$"],
    "transform": [r"^transform\s*logic$", r"^transform$", r"^logic$", r"^cach\s*tinh"],
    "source": [r"^source", r"table\s*/\s*view\s*source", r"^transform\s+source\s+table$",
               r"^nguon"],
}
JOIN_COLS = {
    "name": [r"table\s*/\s*view", r"^table$", r"^bang"],
    "alias": [r"^alias$"],
    "join_type": [r"join\s*type"],
    "condition": [r"^on\b", r"condition"],
    "note": [r"^note$", r"ghi\s*chu"],
}


def _resolve(header: list, spec: dict) -> dict:
    norm = [(h or "").strip().lower() for h in header]
    out = {}
    for key, patterns in spec.items():
        for pat in patterns:
            idx = next((i for i, h in enumerate(norm)
                        if h and re.search(pat, h) and i not in out.values()), None)
            if idx is not None:
                out[key] = idx
                break
    return out


def _at(row: list, cols: dict, key: str) -> str:
    return _cell(row, cols.get(key, -1))
